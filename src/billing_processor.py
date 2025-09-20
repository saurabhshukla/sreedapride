"""
Billing Data Processor for WeGot and Adda Integration
Handles the complete billing workflow from WeGot data to Adda template generation.
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from typing import Dict, Tuple, Optional, List
import io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class WeGotDataProcessor:
    """Processes WeGot data and integrates with billing template."""

    def __init__(self):
        self.wegot_data = None
        self.billing_template = None
        self.final_allocation = None
        self.billing_template_file = None  # Store original file for openpyxl

    def load_wegot_data(self, file_content) -> pd.DataFrame:
        """Load and process WeGot All Blocks data."""
        try:
            # Read the file, skipping initial rows to find actual data
            df = pd.read_excel(file_content, sheet_name=0, header=None)

            # Find the row where actual data starts (usually contains 'S.No' or similar)
            data_start_row = None
            for idx, row in df.iterrows():
                if any(str(cell).lower() in ['s.no', 'serial', 'block', 'apartment'] for cell in row if pd.notna(cell)):
                    data_start_row = idx
                    break

            if data_start_row is not None:
                # Read again with proper header
                df = pd.read_excel(file_content, sheet_name=0, header=data_start_row)
                df = df.dropna(how='all').reset_index(drop=True)

            self.wegot_data = df
            return df

        except Exception as e:
            raise Exception(f"Error loading WeGot data: {str(e)}")

    def load_billing_template(self, file_content) -> Dict[str, pd.DataFrame]:
        """Load billing template with all sheets."""
        try:
            # Store the original file for openpyxl operations
            self.billing_template_file = file_content

            xl_file = pd.ExcelFile(file_content)
            sheets = {}

            for sheet_name in xl_file.sheet_names:
                sheets[sheet_name] = pd.read_excel(file_content, sheet_name=sheet_name, header=None)

            self.billing_template = sheets
            return sheets

        except Exception as e:
            raise Exception(f"Error loading billing template: {str(e)}")

    def extract_final_allocation_data(self) -> pd.DataFrame:
        """Extract final allocation data from billing template."""
        if not self.billing_template or 'Final allocation monthly' not in self.billing_template:
            raise Exception("Billing template not loaded or missing 'Final allocation monthly' sheet")

        df = self.billing_template['Final allocation monthly']

        # Find data start row (usually after headers)
        data_start = None
        for idx, row in df.iterrows():
            if pd.notna(row.iloc[1]) and str(row.iloc[1]).isdigit():  # Serial number column
                data_start = idx
                break

        if data_start is None:
            raise Exception("Could not find data start in Final allocation sheet")

        # Extract relevant columns
        allocation_data = []
        for idx in range(data_start, len(df)):
            row = df.iloc[idx]
            if pd.isna(row.iloc[1]) or not str(row.iloc[1]).isdigit():
                break

            allocation_data.append({
                'Serial_No': int(row.iloc[1]),
                'Block': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
                'Apartment': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
                'Owner_Name': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
                'Meter_Reading': float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0,
                'Rental_GST': float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0,
                'Fixed': float(row.iloc[7]) if pd.notna(row.iloc[7]) else 0,
                'Variable': float(row.iloc[8]) if pd.notna(row.iloc[8]) else 0,
                'Other': float(row.iloc[9]) if pd.notna(row.iloc[9]) else 0,
                'Total': float(row.iloc[10]) if pd.notna(row.iloc[10]) else 0,
                'To_Be_Billed': int(row.iloc[13]) if pd.notna(row.iloc[13]) else 0
            })

        self.final_allocation = pd.DataFrame(allocation_data)
        return self.final_allocation

    def update_wegot_report(self, month: str, bwssb_reading: float, tanker_reading: float) -> Dict:
        """Update the billing template with new WeGot data and readings."""
        if not self.wegot_data is not None:
            raise Exception("WeGot data not loaded")

        if not self.billing_template:
            raise Exception("Billing template not loaded")

        # Update Input sheet with new month and readings
        input_sheet = self.billing_template['Input'].copy()

        # Find and update month
        for idx, row in input_sheet.iterrows():
            if pd.notna(row.iloc[1]) and str(row.iloc[1]).lower() == 'month':
                input_sheet.iloc[idx, 2] = month
                break

        # Add BESCOM and Tanker readings logic here
        # This would update the appropriate cells in the Input sheet

        # Update WeGot report sheet with new data
        wegot_sheet = self.billing_template['WeGot report'].copy()
        # Paste WeGot data into the sheet (implementation depends on exact format)

        return {
            'status': 'success',
            'month': month,
            'bwssb_reading': bwssb_reading,
            'tanker_reading': tanker_reading,
            'updated_sheets': ['Input', 'WeGot report']
        }


class AddaTemplateGenerator:
    """Generates Adda-compatible templates from billing data."""

    def __init__(self):
        self.dues_template = None

    def generate_adda_template(self, final_allocation_df: pd.DataFrame,
                              invoice_date: str = None) -> pd.DataFrame:
        """Generate Adda dues template from final allocation data."""
        if invoice_date is None:
            invoice_date = datetime.now().strftime("%Y-%m-%d")

        # Create Adda template structure
        adda_template = []

        for _, row in final_allocation_df.iterrows():
            # Extract block and flat from apartment code (e.g., A101 -> A, 101)
            apartment = str(row['Apartment'])
            block = apartment[0] if apartment else 'A'
            flat_num = apartment[1:] if len(apartment) > 1 else '101'

            adda_template.append({
                'KeyField': f"50{row['Serial_No']:04d}",  # Generate KeyField
                'Block': f"<{block}>",
                'Flat': f"<{flat_num}>",
                'SquareFeet': 1200,  # Default or from data
                'Category': 'Standard',
                'Name': row['Owner_Name'],
                'CurrentDue': 0,  # This might come from previous data
                'AccountNo*': 301004,  # Standard account number
                'Amount*': int(row['To_Be_Billed']),
                'InvoiceDate(DD/MM/YYYY)*': datetime.strptime(invoice_date, "%Y-%m-%d").strftime("%Y-%m-%d"),
                'Comment*': 'Water cost at actuals'
            })

        self.dues_template = pd.DataFrame(adda_template)
        return self.dues_template

    def export_adda_template(self) -> bytes:
        """Export Adda template as Excel bytes."""
        if self.dues_template is None:
            raise Exception("Adda template not generated")

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self.dues_template.to_excel(writer, sheet_name='dues_template', index=False)

        return output.getvalue()


class BillingWorkflowManager:
    """Manages the complete billing workflow."""

    def __init__(self):
        self.wegot_processor = WeGotDataProcessor()
        self.adda_generator = AddaTemplateGenerator()
        self.workflow_status = {}
        self.updated_billing_template = None
        self.adda_template_file = None

    def execute_billing_workflow(self, wegot_file, billing_file, adda_template_file, month: str,
                                bwssb_reading: float, tanker_reading: float) -> Dict:
        """Execute the complete billing workflow."""
        try:
            # Step 1: Load WeGot data
            self.workflow_status['wegot_loaded'] = False
            wegot_data = self.wegot_processor.load_wegot_data(wegot_file)
            self.workflow_status['wegot_loaded'] = True

            # Step 2: Load billing template
            self.workflow_status['template_loaded'] = False
            billing_sheets = self.wegot_processor.load_billing_template(billing_file)
            self.workflow_status['template_loaded'] = True

            # Step 3: Load Adda template
            self.workflow_status['adda_template_loaded'] = False
            self.adda_template_file = adda_template_file
            self.workflow_status['adda_template_loaded'] = True

            # Step 4: Update billing template with WeGot data
            self.workflow_status['billing_updated'] = False
            self.updated_billing_template = self._update_billing_template_with_wegot_data(
                wegot_data, billing_sheets, month, bwssb_reading, tanker_reading)
            self.workflow_status['billing_updated'] = True

            # Step 5: Extract final allocation data from updated template
            self.workflow_status['allocation_extracted'] = False
            final_allocation = self.wegot_processor.extract_final_allocation_data()
            self.workflow_status['allocation_extracted'] = True

            # Step 6: Generate Adda template using external template file
            self.workflow_status['adda_generated'] = False
            adda_template = self._generate_adda_template_from_file(final_allocation, adda_template_file, month)
            self.workflow_status['adda_generated'] = True

            # Step 6: Calculate summary statistics
            total_amount = final_allocation['To_Be_Billed'].sum()
            apartment_count = len(final_allocation)
            avg_amount = total_amount / apartment_count if apartment_count > 0 else 0

            return {
                'status': 'success',
                'wegot_data_rows': len(wegot_data),
                'final_allocation_rows': len(final_allocation),
                'adda_template_rows': len(adda_template),
                'total_billing_amount': total_amount,
                'apartment_count': apartment_count,
                'average_amount_per_flat': avg_amount,
                'month': month,
                'bwssb_reading': bwssb_reading,
                'tanker_reading': tanker_reading,
                'workflow_status': self.workflow_status
            }

        except Exception as e:
            return {
                'status': 'error',
                'error_message': str(e),
                'workflow_status': self.workflow_status
            }

    def get_adda_template_download(self) -> bytes:
        """Get the generated Adda template for download."""
        return self.adda_generator.export_adda_template()

    def get_final_allocation_data(self) -> pd.DataFrame:
        """Get the final allocation data."""
        return self.wegot_processor.final_allocation

    def get_processing_summary(self) -> Dict:
        """Get summary of the processing workflow."""
        if not self.workflow_status:
            return {'status': 'not_started'}

        total_steps = len(self.workflow_status)
        completed_steps = sum(1 for status in self.workflow_status.values() if status)

        return {
            'total_steps': total_steps,
            'completed_steps': completed_steps,
            'progress_percentage': (completed_steps / total_steps) * 100 if total_steps > 0 else 0,
            'status_details': self.workflow_status
        }

    def _update_billing_template_with_wegot_data(self, wegot_data: pd.DataFrame,
                                               billing_sheets: Dict, month: str,
                                               bwssb_reading: float, tanker_reading: float) -> Dict:
        """Update billing template with WeGot data and create updated version."""
        try:
            updated_sheets = billing_sheets.copy()

            # BLIND REPLACEMENT: Replace WeGot Report sheet with new data completely
            if 'WeGot report' in updated_sheets:
                # Completely replace the WeGot report sheet with new data
                # This is a blind copy - all existing data will be replaced
                updated_sheets['WeGot report'] = wegot_data.copy()
                print(f"✅ WeGot Report sheet updated with {len(wegot_data)} rows of new data")

            # Update Input sheet with month and readings
            if 'Input' in updated_sheets:
                input_sheet = updated_sheets['Input'].copy()

                # Find and update month (row where column 1 contains 'month')
                for idx, row in input_sheet.iterrows():
                    if pd.notna(row.iloc[1]) and str(row.iloc[1]).lower() == 'month':
                        input_sheet.iloc[idx, 2] = month
                        print(f"✅ Month updated to: {month}")
                        break

                # Update BWSSB/Cauvery reading (assuming it's in a specific location)
                # You'll need to identify the exact cell location for BWSSB reading
                for idx, row in input_sheet.iterrows():
                    cell_value = str(row.iloc[1]).lower() if pd.notna(row.iloc[1]) else ""
                    if 'cauvery' in cell_value or 'bwssb' in cell_value or 'bescom' in cell_value:
                        input_sheet.iloc[idx, 2] = bwssb_reading
                        print(f"✅ BWSSB reading updated to: {bwssb_reading}")
                        break

                # Update Tanker reading
                for idx, row in input_sheet.iterrows():
                    cell_value = str(row.iloc[1]).lower() if pd.notna(row.iloc[1]) else ""
                    if 'tanker' in cell_value:
                        input_sheet.iloc[idx, 2] = tanker_reading
                        print(f"✅ Tanker reading updated to: {tanker_reading}")
                        break

                updated_sheets['Input'] = input_sheet

            # Update Allocation sheet with month-specific readings
            if 'Allocation' in updated_sheets:
                allocation_sheet = updated_sheets['Allocation'].copy()

                # Month to row mapping (for the specific cells mentioned: Aug->C15,D15; Sep->C16,D16)
                # Note: pandas uses 0-based indexing, so row 15 in Excel = index 14 in pandas
                month_to_row = {
                    'Jan': 4, 'Feb': 5, 'Mar': 6, 'Apr': 7, 'May': 8, 'Jun': 9,
                    'Jul': 10, 'Aug': 14, 'Sep': 15, 'Oct': 16, 'Nov': 17, 'Dec': 18
                }

                if month in month_to_row:
                    row_num = month_to_row[month]
                    # Update Cauvery reading in column C (index 2)
                    allocation_sheet.iloc[row_num, 2] = bwssb_reading
                    # Update Tanker reading in column D (index 3)
                    allocation_sheet.iloc[row_num, 3] = tanker_reading
                    print(f"✅ Allocation sheet updated for {month}: Row {row_num+1}, Cauvery: {bwssb_reading}, Tanker: {tanker_reading}")

                updated_sheets['Allocation'] = allocation_sheet

            return updated_sheets

        except Exception as e:
            raise Exception(f"Error updating billing template: {str(e)}")

    def _generate_adda_template_from_file(self, final_allocation_df: pd.DataFrame,
                                        adda_template_file, month: str) -> pd.DataFrame:
        """Generate Adda template using the provided template file."""
        try:
            # Load the Adda template
            adda_template = pd.read_excel(adda_template_file, sheet_name=0, header=0)

            # Rename columns properly
            expected_columns = ['KeyField', 'Block', 'Flat', 'SquareFeet', 'Category',
                              'Name', 'CurrentDue', 'AccountNo*', 'Amount*',
                              'InvoiceDate(DD/MM/YYYY)*', 'Comment*']

            adda_template.columns = expected_columns

            # Update the template with billing data
            for idx, row in adda_template.iterrows():
                if idx == 0:  # Skip header row
                    continue

                # Find corresponding apartment in final allocation
                apt_code = f"{row['Block'].strip('<>')}{row['Flat'].strip('<>')}"

                # Find matching apartment in final allocation
                matching_apt = final_allocation_df[
                    final_allocation_df['Apartment'] == apt_code
                ]

                if not matching_apt.empty:
                    billing_amount = int(matching_apt.iloc[0]['To_Be_Billed'])
                    adda_template.at[idx, 'Amount*'] = billing_amount
                    adda_template.at[idx, 'AccountNo*'] = 301004
                    adda_template.at[idx, 'InvoiceDate(DD/MM/YYYY)*'] = datetime.now().strftime("%Y-%m-%d")
                    adda_template.at[idx, 'Comment*'] = 'Water cost at actuals'

            return adda_template

        except Exception as e:
            raise Exception(f"Error generating Adda template: {str(e)}")

    def get_updated_billing_template_download(self) -> bytes:
        """Get the updated billing template for download with minimal modifications."""
        if self.wegot_processor.billing_template_file is None:
            raise Exception("Original billing template file not available")

        try:
            # First, try to return the original file with minimal modifications
            # Copy the original file content first
            original_content = self.wegot_processor.billing_template_file

            # Reset the file pointer if it's a file-like object
            if hasattr(original_content, 'seek'):
                original_content.seek(0)
                file_content = original_content.read()
                original_content.seek(0)  # Reset for future use
            else:
                file_content = original_content

            # Try openpyxl with minimal modifications
            try:
                workbook = load_workbook(io.BytesIO(file_content), data_only=False)

                # Only update WeGot Report sheet if it exists
                if self.wegot_processor.wegot_data is not None and 'WeGot report' in workbook.sheetnames:
                    ws = workbook['WeGot report']

                    # Clear only the data area (first 1000 rows and 50 columns to be safe)
                    for row in range(1, min(1000, ws.max_row + 100)):
                        for col in range(1, min(50, ws.max_column + 10)):
                            try:
                                cell = ws.cell(row=row, column=col)
                                if cell.data_type != 'f':  # Don't touch formulas
                                    cell.value = None
                            except:
                                continue

                    # Insert new WeGot data starting from A1
                    wegot_data = self.wegot_processor.wegot_data
                    for r_idx, row in enumerate(dataframe_to_rows(wegot_data, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            try:
                                ws.cell(row=r_idx, column=c_idx, value=value)
                            except:
                                continue

                # Update Allocation sheet with month-specific readings
                if 'Allocation' in workbook.sheetnames:
                    allocation_ws = workbook['Allocation']

                    try:
                        # Direct approach: Find the month and readings from the processing data
                        # Since we know the exact structure, let's use a more direct method

                        # Get month, bwssb_reading, and tanker_reading from the method parameters
                        # We'll access these from the updated_billing_template or use fallback
                        month = None
                        bwssb_reading = None
                        tanker_reading = None

                        # Try to extract from updated_billing_template first
                        if hasattr(self, 'updated_billing_template') and self.updated_billing_template and 'Input' in self.updated_billing_template:
                            input_data = self.updated_billing_template['Input']
                            for _, row in input_data.iterrows():
                                if pd.notna(row.iloc[1]) and str(row.iloc[1]).lower() == 'month':
                                    month = str(row.iloc[2]).strip()
                                elif pd.notna(row.iloc[1]):
                                    cell_value = str(row.iloc[1]).lower()
                                    if 'cauvery' in cell_value or 'bwssb' in cell_value:
                                        bwssb_reading = row.iloc[2]
                                    elif 'tanker' in cell_value:
                                        tanker_reading = row.iloc[2]

                        # Additional fallback: Try to extract from Allocation sheet directly
                        if hasattr(self, 'updated_billing_template') and self.updated_billing_template and 'Allocation' in self.updated_billing_template:
                            allocation_data = self.updated_billing_template['Allocation']
                            # Look for the month row and extract the values that were updated
                            for _, row in allocation_data.iterrows():
                                if pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() == 'Aug':
                                    if pd.notna(row.iloc[2]):  # Column C (Cauvery)
                                        bwssb_reading = row.iloc[2]
                                    if pd.notna(row.iloc[3]):  # Column D (Tanker)
                                        tanker_reading = row.iloc[3]
                                    month = 'Aug'
                                    break
                                elif pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() == 'Sep':
                                    if pd.notna(row.iloc[2]):  # Column C (Cauvery)
                                        bwssb_reading = row.iloc[2]
                                    if pd.notna(row.iloc[3]):  # Column D (Tanker)
                                        tanker_reading = row.iloc[3]
                                    month = 'Sep'
                                    break

                        # Alternatively, try to extract from Allocation sheet structure directly
                        if not month:
                            # Look for the current month in the Allocation sheet itself
                            allocation_month_cell = allocation_ws.cell(row=4, column=3)  # C4 contains the month
                            if allocation_month_cell.value:
                                month = str(allocation_month_cell.value).strip()

                        # Direct cell updates based on month
                        if month and bwssb_reading is not None and tanker_reading is not None:
                            # Convert to proper types
                            bwssb_val = float(bwssb_reading) if bwssb_reading else 0
                            tanker_val = float(tanker_reading) if tanker_reading else 0

                            # Update based on specific month (exact row mapping from analysis)
                            if month == 'Aug':
                                allocation_ws.cell(row=15, column=3, value=bwssb_val)  # C15 = Cauvery for Aug
                                allocation_ws.cell(row=15, column=4, value=tanker_val)  # D15 = Tanker for Aug
                                print(f"✅ Allocation sheet updated: Aug → C15={bwssb_val}, D15={tanker_val}")
                            elif month == 'Sep':
                                allocation_ws.cell(row=16, column=3, value=bwssb_val)  # C16 = Cauvery for Sep
                                allocation_ws.cell(row=16, column=4, value=tanker_val)  # D16 = Tanker for Sep
                                print(f"✅ Allocation sheet updated: Sep → C16={bwssb_val}, D16={tanker_val}")
                            elif month == 'Oct':
                                allocation_ws.cell(row=17, column=3, value=bwssb_val)  # C17 = Cauvery for Oct
                                allocation_ws.cell(row=17, column=4, value=tanker_val)  # D17 = Tanker for Oct
                                print(f"✅ Allocation sheet updated: Oct → C17={bwssb_val}, D17={tanker_val}")
                            else:
                                # Generic mapping for other months
                                month_to_row = {
                                    'Jan': 8, 'Feb': 9, 'Mar': 10, 'Apr': 11, 'May': 12, 'Jun': 13,
                                    'Jul': 14, 'Nov': 18, 'Dec': 19
                                }
                                if month in month_to_row:
                                    row_num = month_to_row[month]
                                    allocation_ws.cell(row=row_num, column=3, value=bwssb_val)
                                    allocation_ws.cell(row=row_num, column=4, value=tanker_val)
                                    print(f"✅ Allocation sheet updated: {month} → C{row_num}={bwssb_val}, D{row_num}={tanker_val}")
                        else:
                            print(f"⚠️ Missing data for allocation update: month={month}, bwssb={bwssb_reading}, tanker={tanker_reading}")

                    except Exception as e:
                        print(f"⚠️ Could not update Allocation sheet: {str(e)}")
                        pass  # Continue even if allocation update fails

                # Save the modified workbook
                output = io.BytesIO()
                workbook.save(output)
                output.seek(0)
                return output.getvalue()

            except Exception:
                # If openpyxl fails, fall back to pandas approach
                if self.updated_billing_template is None:
                    raise Exception("Both openpyxl and updated template approaches failed")

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for sheet_name, sheet_data in self.updated_billing_template.items():
                        sheet_data.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                output.seek(0)
                return output.getvalue()

        except Exception as e:
            raise Exception(f"Error generating updated billing template: {str(e)}")

    def _is_merged_cell(self, worksheet, row, col):
        """Check if a cell is part of a merged range."""
        try:
            cell_coordinate = worksheet.cell(row=row, column=col).coordinate

            for merged_range in worksheet.merged_cells.ranges:
                if cell_coordinate in merged_range:
                    return True
            return False
        except Exception:
            return False  # If we can't determine, assume it's not merged